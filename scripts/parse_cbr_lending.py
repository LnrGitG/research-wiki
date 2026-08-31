#!/usr/bin/env python3
"""Parse CBR lending XLSX files into SQLite cbr_lending.db."""
import re, sqlite3, glob, os
from openpyxl import load_workbook

RAW = "/home/lnr/research-wiki/raw/cbr"
DB = "/home/lnr/research-wiki/data/cbr_lending.db"

MONTHS = {"январь":1,"февраль":2,"март":3,"апрель":4,"май":5,"июнь":6,"июль":7,
          "август":8,"сентябрь":9,"октябрь":10,"ноябрь":11,"декабрь":12}

def norm(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace("\u00a0","").replace(" ","")
    s = s.replace(",", ".")
    if s in ("", "-", "—", "…", "x"): return None
    try: return float(s)
    except ValueError: return None

def month_from_label(label):
    m = re.search(r"([а-яё]+)\s*(\d{4})", str(label).lower())
    if m and m.group(1) in MONTHS:
        return f"{int(m.group(2))}-{MONTHS[m.group(1)]:02d}-01"
    return None

def month_from_header_cell(cell):
    """Header may be datetime or 'Июнь 2025'."""
    import datetime
    if isinstance(cell, (datetime.datetime, datetime.date)):
        return f"{cell.year}-{cell.month:02d}-01"
    return month_from_label(cell)

def init_db(con):
    con.executescript("""
    CREATE TABLE IF NOT EXISTS mortgage_monthly(region_name TEXT, report_date TEXT, indicator TEXT, value REAL, unit TEXT, PRIMARY KEY(region_name,report_date,indicator));
    CREATE TABLE IF NOT EXISTS escrow_monthly(region_name TEXT, report_date TEXT, indicator TEXT, value REAL, unit TEXT, PRIMARY KEY(region_name,report_date,indicator));
    CREATE TABLE IF NOT EXISTS corporate_monthly(region_name TEXT, report_date TEXT, indicator TEXT, value REAL, unit TEXT, PRIMARY KEY(region_name,report_date,indicator));
    """)

con = sqlite3.connect(DB)
init_db(con)

# ---------- 1. VFS corporate & mortgage files ----------
VFS = [
    ("vfs/01_04_D_New_loans_subj.xlsx", "corporate", "Новые кредиты юридическим лицам", "млн руб", "corporate_monthly"),
    ("vfs/01_05_D_Debt_subj.xlsx",      "corporate", "Задолженность юридических лиц", "млн руб", "corporate_monthly"),
    ("vfs/02_04_New_loans_ind.xlsx",    "mortgage",  "Новые ипотечные кредиты физлицам", "млн руб", "mortgage_monthly"),
    ("vfs/02_05_Debt_ind.xlsx",         "mortgage",  "Задолженность по кредитам физлицам", "млн руб", "mortgage_monthly"),
]
for path, kind, base, unit, table in VFS:
    wb = load_workbook(os.path.join(RAW, path), data_only=True, read_only=True)
    rows_written = 0
    for ws in wb.worksheets:
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        # header row detection: first row where >=3 cells parse as months
        hidx = None
        for i in range(min(6, len(grid))):
            if sum(1 for c in grid[i] if month_from_header_cell(c)) >= 3:
                hidx = i; break
        if hidx is None: continue
        headers = grid[hidx]
        # columns with month dates
        col_month = []
        for j, h in enumerate(headers):
            m = month_from_header_cell(h)
            if m: col_month.append((j, m))
        if not col_month: continue
        for row in grid[hidx + 1:]:
            name = row[0] if row else None
            if not name or not isinstance(name, str): continue
            name = name.strip()
            if not name or "ФЕДЕРАЛЬНЫЙ ОКРУГ" in name.upper(): continue
            for j, m in col_month:
                v = norm(row[j]) if j < len(row) else None
                if v is not None:
                    ind = f"{base} ({ws.title})"
                    con.execute(f"INSERT OR REPLACE INTO {table} VALUES (?,?,?,?,?)", (name, m, ind, v, unit))
                    rows_written += 1
    print(f"{path}: {rows_written} rows")
    wb.close()

# ---------- 2. Escrow files ----------
esc_files = sorted(glob.glob(os.path.join(RAW, "escrow", "*.xlsx")))
def parse_date(fn):
    b = os.path.basename(fn)
    d, m, y = b[:2], b[2:4], b[4:8]
    return f"{y}-{m}-{d}"

IND_COLS = {
    2: ("Кол-во действующих кредитных договоров", "шт"),
    3: ("Сумма действующих кредитных договоров", "млн руб"),
    4: ("Задолженность", "млн руб"),
    5: ("Кол-во счетов эскроу", "шт"),
    6: ("Кол-во счетов эскроу с остатком", "шт"),
    7: ("Остатки средств на счетах эскроу", "млн руб"),
    8: ("Средневзвешенная ставка по кредитам", "%"),
    9: ("Кол-во раскрытых счетов эскроу", "шт"),
    10: ("Сумма средств с раскрытых счетов эскроу", "млн руб"),
}
tot = 0
for f in esc_files:
    rd = parse_date(f)
    wb = load_workbook(f, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if not name or name.lower() in ("итого",) or name.isdigit(): continue
        for j, (ind, unit) in IND_COLS.items():
            v = norm(row[j]) if j < len(row) else None
            if v is not None:
                con.execute("INSERT OR REPLACE INTO escrow_monthly VALUES (?,?,?,?,?)", (name, rd, ind, v, unit))
                tot += 1
    wb.close()
print(f"escrow: {tot} rows from {len(esc_files)} files")

# ---------- 3. Mortgage bulletin Section II (Т_16..Т_34) ----------
bl = os.path.join(RAW, "bulletin", "mortgage_lending_market_2606_74.xlsx")
wb = load_workbook(bl, data_only=True, read_only=True)
INDICATORS = {
    "Т_16": ("Количество ИЖК", "шт"), "Т_17": ("Количество ИЖК по ДДУ", "шт"),
    "Т_18": ("Количество ИЖК на цели ИЖС", "шт"), "Т_19": ("Объём ИЖК", "млн руб"),
    "Т_20": ("Объём ИЖК по ДДУ", "млн руб"), "Т_21": ("Объём ИЖК на цели ИЖС", "млн руб"),
    "Т_22": ("Задолженность по ИЖК", "млн руб"), "Т_23": ("Задолженность по ИЖК по ДДУ", "млн руб"),
    "Т_24": ("Просроченная задолженность по ИЖК", "млн руб"), "Т_25": ("Просроченная задолженность по ИЖК по ДДУ", "млн руб"),
    "Т_26": ("Ставка по ИЖК", "%"), "Т_27": ("Ставка по ИЖК по ДДУ", "%"),
    "Т_28": ("Ставка по ИЖК без учёта ДДУ", "%"), "Т_29": ("Ставка по ИЖК на цели ИЖС", "%"),
    "Т_30": ("Доля ИЖК в общем объёме кредитов", "%"), "Т_31": ("Доля задолженности по ИЖК", "%"),
    "Т_32": ("Средний размер ИЖК", "млн руб"), "Т_33": ("Средний размер ИЖК по ДДУ", "млн руб"),
    "Т_34": ("Средняя цена 1 кв.м", "тыс руб"),
}
cnt = 0
for ws in wb.worksheets:
    key = ws.title.strip().replace("T_", "Т_")
    if key not in INDICATORS: continue
    ind, unit = INDICATORS[key]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    if len(grid) < 5: continue
    headers = grid[3]  # row 4 = month headers
    col_month = [(j, month_from_header_cell(h)) for j, h in enumerate(headers)]
    col_month = [(j, m) for j, m in col_month if m]
    if not col_month: continue
    for row in grid[4:]:
        name = row[0] if row else None
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if not name or name.endswith("ОКРУГ"): continue  # skip federal districts
        for j, m in col_month:
            v = norm(row[j]) if j < len(row) else None
            if v is not None:
                con.execute("INSERT OR REPLACE INTO mortgage_monthly VALUES (?,?,?,?,?)", (name, m, ind, v, unit))
                cnt += 1
print(f"mortgage bulletin: {cnt} rows")
wb.close()
con.commit()

# ---------- Summary ----------
for t in ("mortgage_monthly", "escrow_monthly", "corporate_monthly"):
    n, latest = con.execute(f"SELECT COUNT(*), MAX(report_date) FROM {t}").fetchone()
    print(f"{t}: {n} records, latest {latest}")

print("\nTop 5 regions by mortgage volume (latest bulletin month, Т_19 Объём ИЖК):")
lates = con.execute("SELECT MAX(report_date) FROM mortgage_monthly WHERE indicator='Объём ИЖК'").fetchone()[0]
for r in con.execute("SELECT region_name, value FROM mortgage_monthly WHERE indicator='Объём ИЖК' AND report_date=? AND region_name NOT IN ('Российская Федерация') ORDER BY value DESC LIMIT 5", (lates,)):
    print(f"  {r[0]}: {r[1]:,.0f} млн руб ({lates})")
con.close()
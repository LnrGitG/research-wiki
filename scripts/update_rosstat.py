#!/usr/bin/env python3
"""
Скрипт обновления локальной базы данных Росстат (Строительство).

База: data/rosstat_construction.db (SQLite)
Источник: https://rosstat.gov.ru/folder/14458

Использование:
    python scripts/update_rosstat.py              # обновить все файлы
    python scripts/update_rosstat.py --check        # проверить наличие новых файлов
    python scripts/update_rosstat.py --rebuild      # пересоздать базу с нуля
    python scripts/update_rosstat.py --download-only # только скачать, не парсить
"""

import os, sys, re, sqlite3, subprocess, argparse
from datetime import datetime
import openpyxl, xlrd

# Paths
REPO_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(REPO_DIR, 'raw', 'rosstat', 'data')
DB_PATH = os.path.join(REPO_DIR, 'data', 'rosstat_construction.db')
BASE_URL = 'https://rosstat.gov.ru/storage/mediabank'

# File catalog: filename → (category, description, priority)
FILE_CATALOG = {
    # ★★★ Priority — housing input
    'vv_jil_dom-graf_07-2026.xls': ('housing_input_operational', 'Ввод жилых домов (график)', 3),
    'jil_dom-oper_07-2026.xls': ('housing_input_operational', 'Жилые дома по субъектам (оперативные)', 3),
    'stroi_131_2025.xls': ('housing_input_annual', 'Ввод жилых домов в РФ (годовой)', 3),
    'stroi_135.xls': ('housing_cost', 'Средняя стоимость строительства', 3),
    'stroi_135-sub.xls': ('housing_cost', 'Стоимость по субъектам', 3),
    'vvod_jil_dom_RF_2025.xls': ('housing_input_regions', 'Ввод жилых домов по регионам', 3),
    
    # ★★☆ High
    'stroi_134_2025.xls': ('housing_construction', 'Число построенных квартир', 2),
    'Raboti_stroi_2025.xlsx': ('construction_works', 'Объём работ (СМР)', 2),
    'stroi_et_2025.xlsx': ('housing_input_by_floor', 'По этажности', 2),
    'stroi_sten_2025.xls': ('housing_input_by_walls', 'По материалам стен', 2),
    'stroi_136_2025.xls': ('unfinished_construction', 'Незавершённое строительство', 2),
    'Stroi-1000_2025.xlsx': ('housing_per_1000', 'На 1000 чел. населения', 2),
    
    # ★☆☆ Medium
    'str-stoim.xls': ('construction_cost', 'Стоимость строительства', 1),
    'Zatrat_stroi.xlsx': ('cost_structure', 'Структура затрат', 1),
    'Stroi_111_2025.xls': ('buildings_input', 'Ввод зданий', 1),
    'Stroi_112_2025.xls': ('nonresidential_buildings', 'Нежилые здания', 1),
    'stroi_121_2025.xlsx': ('production_capacity', 'Производственные мощности', 1),
    'stroi_141_2025.xls': ('social_facilities_education', 'Объекты образования', 1),
    'Stroi_142_2025.xls': ('social_facilities_health', 'Объекты здравоохранения', 1),
    'Stroi_143_2025.xls': ('social_facilities_culture', 'Объекты культуры', 1),
    'Stroi_144_2025.xls': ('social_facilities_utilities', 'Объекты ЖКХ', 1),
    'Stroi_151_2025.xls': ('unfinished_construction', 'Незавершённое (число)', 1),
    'Nezaversh_stroi_RF_2025.xls': ('unfinished_construction', 'Незавершённое (детали)', 1),
    'Stroi_3_2025.xls': ('unfinished_construction', 'Незавершённое (регионы)', 1),
    'stroi_ob_1990.xlsx': ('construction_organizations', 'Деятельность организаций (с 1990)', 1),
    'Stroi_sub.xlsx': ('construction_works', 'Объём работ по субъектам', 1),
    'Stroi_form_proc.xlsx': ('construction_by_ownership', 'По формам собственности', 1),
    'str_mosh_2kv_2026.xlsx': ('capacity_utilization', 'Использование мощностей', 1),
    
    # Operational 2026
    'vv-mosh-oper_2kv-2026.xlsx': ('capacity_operational', 'Мощности (оперативные)', 2),
    'vv-mosh-sub_2kv-2026.xlsx': ('capacity_operational', 'Мощности по субъектам', 2),
    'vv-zd-oper_2kv-2026.xlsx': ('buildings_operational', 'Здания (оперативные)', 2),
    'vv-zd-sub_07-2026.xlsx': ('buildings_operational', 'Здания по субъектам', 2),
    'vv-sockul-oper_2kv-2026.xlsx': ('social_operational', 'Соцкультобъекты', 2),
    'vv-sockul-sub_2kv-2026.xlsx': ('social_operational', 'Соцкультобъекты по субъектам', 2),
    'Operativ_06.xlsx': ('operational_summary', 'Сводные оперативные данные', 2),
}


def download_file(filename, force=False):
    """Download a file from Rosstat. Returns True if downloaded or already exists."""
    out_path = os.path.join(DATA_DIR, filename)
    
    if os.path.exists(out_path) and not force:
        return True, 'exists'
    
    url = f"{BASE_URL}/{filename}"
    result = subprocess.run(
        ['curl', '-sk', '-L', '-o', out_path, url],
        capture_output=True, timeout=60
    )
    
    if os.path.exists(out_path) and os.path.getsize(out_path) > 100:
        return True, 'downloaded'
    return False, 'failed'


def detect_period(s):
    s = str(s).strip() if s is not None else ''
    m = re.match(r'^(19\d{2}|20\d{2})(\.0)?$', s)
    if m:
        return m.group(1), 'year'
    m = re.match(r'^(19\d{2})-(19\d{2}|20\d{2})$', s)
    if m:
        return s, 'year_range'
    months = {'январь':1, 'февраль':2, 'март':3, 'апрель':4, 'май':5, 'июнь':6,
              'июль':7, 'август':8, 'сентябрь':9, 'октябрь':10, 'ноябрь':11, 'декабрь':12}
    for mname, mnum in months.items():
        if s.lower().startswith(mname) and '20' in s:
            year_m = re.search(r'20\d{2}', s)
            if year_m:
                return f"{year_m.group()}-{mnum:02d}", 'month'
    for mname in months:
        if mname in s.lower() and '20' in s:
            year_m = re.search(r'20\d{2}', s)
            if year_m:
                return s, 'cumulative'
    return None, None


def clean_value(v):
    if v is None or v == '':
        return None, None
    if isinstance(v, (int, float)):
        return float(v), str(v)
    s = str(v).strip().replace(',', '.').replace('\xa0', '').replace(' ', '')
    if s in ('-', '', '–', '—', '…'):
        return None, s
    try:
        return float(s), s
    except:
        return None, str(v)


def clean_text(s):
    if s is None:
        return ''
    s = str(s).replace('\xa0', ' ').replace('\n', ' ').strip()
    return re.sub(r'\s+', ' ', s)


def get_rows(path, ext, sheet_name):
    if ext == 'xlsx':
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows
    else:
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name(sheet_name)
        rows = [[ws.cell_value(i, j) for j in range(ws.ncols)] for i in range(ws.nrows)]
        del wb
        return rows


def get_sheet_names(path, ext):
    if ext == 'xlsx':
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    else:
        wb = xlrd.open_workbook(path)
        names = wb.sheet_names()
        del wb
        return names


def parse_file(path, filename, category, conn, updated_at):
    """Parse a single file and insert into database."""
    c = conn.cursor()
    ext = filename.split('.')[-1].lower()
    obs_count = 0
    
    try:
        sheets = get_sheet_names(path, ext)
    except Exception as e:
        return 0, str(e)
    
    for sn in sheets:
        if 'содержание' in sn.lower():
            continue
        
        try:
            rows = get_rows(path, ext, sn)
        except:
            continue
        
        if len(rows) < 3:
            continue
        
        # Find header row
        header_row_idx = None
        header_periods = []
        for i, row in enumerate(rows[:15]):
            periods = [(detect_period(cell)) for cell in row]
            valid = [p for p, pt in periods if p and pt == 'year']
            if len(valid) >= 3:
                header_row_idx = i
                header_periods = periods
                break
        
        if header_row_idx is None:
            for i, row in enumerate(rows[:10]):
                periods = [(detect_period(cell)) for cell in row]
                valid = [p for p, pt in periods if p]
                if len(valid) >= 2:
                    header_row_idx = i
                    header_periods = periods
                    break
        
        if header_row_idx is None:
            continue
        
        indicator_name = sn
        for i in range(min(header_row_idx, 10)):
            text = clean_text(rows[i][0]) if rows[i] else ''
            if text and text != 'К содержанию' and len(text) > 5:
                indicator_name = text
                break
        
        unit = ''
        for i in range(min(header_row_idx + 2, len(rows))):
            for cell in rows[i][:3]:
                text = clean_text(cell)
                if text and '(' in text and ')' in text:
                    unit = text
                    break
            if unit:
                break
        
        c.execute('''
            INSERT INTO indicators (indicator_name, unit, category, source_file, sheet_name, description)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (indicator_name[:200], unit[:100], category, filename, sn,
              f"Раздел: {category}; Лист: {sn}"))
        indicator_id = c.lastrowid
        
        data_start = header_row_idx + 1
        for i in range(data_start, min(data_start + 500, len(rows))):
            row = rows[i]
            if not row:
                continue
            
            row_label = clean_text(row[0])
            if not row_label or row_label in ('К содержанию', 'А'):
                continue
            
            has_data = any(clean_value(cell)[0] is not None for cell in row[1:])
            if not has_data:
                continue
            
            for j, cell in enumerate(row[1:], 1):
                if j - 1 < len(header_periods):
                    p, pt = header_periods[j - 1]
                    if not p:
                        continue
                    val, val_str = clean_value(cell)
                    if val is None and not val_str:
                        continue
                    c.execute('''
                        INSERT INTO observations (indicator_id, region_name, row_label, period, period_type, value, value_str, source_file, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (indicator_id, row_label[:200], row_label[:200], p, pt, val, val_str, filename, updated_at))
                    obs_count += 1
    
    return obs_count, None


def create_schema(conn):
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS indicators (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            indicator_name TEXT NOT NULL,
            unit TEXT,
            category TEXT,
            source_file TEXT,
            sheet_name TEXT,
            description TEXT
        );
        CREATE TABLE IF NOT EXISTS observations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            indicator_id INTEGER NOT NULL,
            region_name TEXT,
            row_label TEXT,
            period TEXT NOT NULL,
            period_type TEXT,
            value REAL,
            value_str TEXT,
            source_file TEXT,
            updated_at TEXT,
            FOREIGN KEY (indicator_id) REFERENCES indicators(id)
        );
        CREATE INDEX IF NOT EXISTS idx_obs_indicator ON observations(indicator_id);
        CREATE INDEX IF NOT EXISTS idx_obs_period ON observations(period);
        CREATE INDEX IF NOT EXISTS idx_obs_region ON observations(region_name);
        CREATE TABLE IF NOT EXISTS update_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_at TEXT NOT NULL,
            action TEXT NOT NULL,
            files_processed INTEGER,
            indicators_added INTEGER,
            observations_added INTEGER,
            notes TEXT
        );
    ''')
    conn.commit()


def main():
    parser = argparse.ArgumentParser(description='Обновление базы Росстат (Строительство)')
    parser.add_argument('--check', action='store_true', help='Проверить наличие новых файлов')
    parser.add_argument('--rebuild', action='store_true', help='Пересоздать базу с нуля')
    parser.add_argument('--download-only', action='store_true', help='Только скачать файлы')
    parser.add_argument('--priority', type=int, choices=[1, 2, 3], 
                        help='Скачать только файлы с приоритетом ≥ N')
    args = parser.parse_args()
    
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    
    updated_at = datetime.now().isoformat()
    
    # Step 1: Download
    print("=== Шаг 1: Загрузка файлов ===")
    downloaded = 0
    skipped = 0
    failed = 0
    
    for filename, (category, desc, priority) in sorted(FILE_CATALOG.items()):
        if args.priority and priority < args.priority:
            continue
        
        ok, status = download_file(filename, force=args.rebuild)
        if status == 'downloaded':
            print(f"  ⬇  {filename} — {desc}")
            downloaded += 1
        elif status == 'exists':
            skipped += 1
        else:
            print(f"  ❌ {filename} — failed")
            failed += 1
    
    print(f"\nИтого: {downloaded} скачано, {skipped} уже есть, {failed} ошибок")
    
    if args.download_only:
        return
    
    if args.check:
        print("\n=== Проверка (без записи в БД) ===")
        for filename in sorted(FILE_CATALOG.keys()):
            path = os.path.join(DATA_DIR, filename)
            if os.path.exists(path):
                size = os.path.getsize(path)
                print(f"  ✅ {filename}: {size:,} bytes")
            else:
                print(f"  ❌ {filename}: не загружен")
        return
    
    # Step 2: Parse and build database
    print("\n=== Шаг 2: Парсинг и загрузка в БД ===")
    
    if args.rebuild and os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    
    conn = sqlite3.connect(DB_PATH)
    create_schema(conn)
    
    if args.rebuild:
        # Clear existing data
        conn.execute('DELETE FROM observations')
        conn.execute('DELETE FROM indicators')
        conn.commit()
    
    total_obs = 0
    total_ind = 0
    files_processed = 0
    
    for filename, (category, desc, priority) in sorted(FILE_CATALOG.items()):
        if args.priority and priority < args.priority:
            continue
        
        path = os.path.join(DATA_DIR, filename)
        if not os.path.exists(path):
            continue
        
        obs_count, err = parse_file(path, filename, category, conn, updated_at)
        if err:
            print(f"  ❌ {filename}: {err}")
        else:
            if obs_count > 0:
                print(f"  ✅ {filename}: {obs_count} набл.")
            total_obs += obs_count
            files_processed += 1
    
    conn.commit()
    
    # Log
    conn.execute('''
        INSERT INTO update_log (run_at, action, files_processed, observations_added, notes)
        VALUES (?, ?, ?, ?, ?)
    ''', (updated_at, 'rebuild' if args.rebuild else 'update', files_processed, total_obs,
          f"Priority >= {args.priority}" if args.priority else "all"))
    conn.commit()
    
    # Stats
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM indicators')
    total_ind = c.fetchone()[0]
    c.execute('SELECT COUNT(*) FROM observations')
    total_obs_db = c.fetchone()[0]
    
    print(f"\n{'='*50}")
    print(f"Файлов обработано: {files_processed}")
    print(f"Индикаторов в БД: {total_ind}")
    print(f"Наблюдений в БД: {total_obs_db}")
    print(f"Новых наблюдений: {total_obs}")
    print(f"База: {os.path.getsize(DB_PATH):,} bytes ({os.path.getsize(DB_PATH)/1024/1024:.1f} MB)")
    
    c.execute('SELECT category, COUNT(DISTINCT i.id), COUNT(o.id) FROM indicators i LEFT JOIN observations o ON i.id = o.indicator_id GROUP BY category ORDER BY COUNT(o.id) DESC')
    print(f"\n=== По категориям ===")
    for cat, ind_cnt, obs_cnt in c.fetchall():
        print(f"  {cat}: {ind_cnt} инд., {obs_cnt} набл.")
    
    c.execute('SELECT MIN(period), MAX(period) FROM observations WHERE period_type = "year"')
    min_p, max_p = c.fetchone()
    print(f"\nПериод (годовые): {min_p} — {max_p}")
    
    conn.close()
    print(f"\n✅ Готово. База: {DB_PATH}")


if __name__ == '__main__':
    main()